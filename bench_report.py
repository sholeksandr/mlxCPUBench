#!/usr/local/bin/python2.7

# pylint: disable=old-style-class, no-init, missing-docstring, line-too-long
# pylint: disable=too-few-public-methods
# pylint: disable= C0103

'''
Created on Dec 05, 2019

Author: Oleksandr Shamray <oleksandrs@mellanox.com>
Version: 1.0

Description:

'''


#############################
# Global imports
#############################
import sys
import json
import os.path
from datetime import datetime

import pdb

try:
    xls_format_ena = True
    # import openpyxl module
    import openpyxl
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    xls_format_ena = False
    # module doesn't exist, deal with it.


#############################
# Global const
#############################
# name of benchmark summary file
BENCHMARK_SUMMARY_FILENAME = 'benchmark_summary'

# delete temporary files after generate report succesfully finished
CPU_USAGE_TEMP_CLEANUP = True

# format of summary test results in benchmark summary file
TEST_REPORT_SUMARY_FORMAT = "Summary: Test count: {count}, passed: {passed}, Tests duration: {total_time} sec, Overal tests duration: {total_time_raw} sec"

# 24h time format
TIME_FORMAT_24 = "%H:%M:%S"
# 12h time format
TIME_FORMAT_12 = "%I:%M:%S %p"

# date-time format of test results in benchmark summary file
TEST_REPORT_DATETIME_FORMAT = "%H:%M:%S"

# format of test results in benchmark summary file
TEST_REPORT_FORMAT = "Test{test_num} [{test_name}] {test_iter}: start={start}; end={end}; overal_duration={overal_time}, test_duration={test_time} {extra_info}"

#############################
# Global variables
#############################


def check_header(line):
    '''
    @summary: check if cpu usage header is lin line
    @param line: string line to check
    @return True if header found
    '''
    return '%iowait' in line


def get_time(time_str):
    '''
    @summary: parce time string into datetime obj
    @param time_str:  string with time
    @return datetime obj
    '''
    if time_str[-2:] in ["AM", "PM", "am", "pm"]:
        return datetime.strptime(time_str, TIME_FORMAT_12)
    else:
        return datetime.strptime(time_str, TIME_FORMAT_24)


def parse_cpu_log(test_res):
    '''
    @summary: parce cpu usage files and save it into dict
    @param test_res: tesr tresults list
    @return dict
    '''
    start = None
    header_found = False
    results = []
    record = {}
    start_time = datetime(1900, 1, 1, 0, 0, 0)
    if test_res['cpu_log_align']:
        start_time = test_res['start_time']
    start_time = start_time.replace(year=1900, month=1, day=1)

    with open(test_res['cpu_log'], "r") as log_file:
        for line in log_file:
            if check_header(line):
                
                line = line.split('  ')
                line = filter(None, line)
                time = get_time(line[0])
                if time < start_time:
                    continue
                
                header_found = True               
                if not start:
                    start = time
                time = time - start
                record = {'time': time.total_seconds(), 'CPU': {}}
            elif ((line == '') or (line == '\n')) and header_found:
                results.append(record)
                header_found = False
            elif header_found:
                line = line.split('  ')
                line = filter(None, line)
                cpu_stat = {}
                cpu_stat['usr'] = float(line[2])
                cpu_stat['nice'] = float(line[3])
                cpu_stat['sys'] = float(line[4])
                cpu_stat['iowait'] = float(line[5])
                cpu_stat['irq'] = float(line[6])
                cpu_stat['soft'] = float(line[7])
                cpu_stat['steal'] = float(line[8])
                cpu_stat['guest'] = float(line[9])
                cpu_stat['gnice'] = float(line[10])
                cpu_stat['idle'] = float(line[11].rstrip(os.linesep))
                record['CPU'][line[1]] = cpu_stat
    if record:
        results.append(record)
    return results


def format_cpu_csv_output(cpu_res_arr, output_file):
    '''
    @summary: process CPU load results and save it to csv file
    @param cpu_res_arr: dict with parced cpu load test results
    @param output_file: output file name to save results
    @return True on successfull
    '''
    if not cpu_res_arr:
        return True;
    output_file_name = output_file + '.csv'
    output_file = open(output_file_name, "w")
    print_header = False
    for stat_rec in cpu_res_arr:
        if not print_header:
            core_list = sorted(stat_rec['CPU'].keys())
            header = ['time'] + core_list
            output_file.write(','.join(header) + '\n')
            print_header = True
        stat_line = [str(stat_rec['time'])]
        for core in core_list:
            core_stat = stat_rec['CPU'][core]
            usage = core_stat['usr'] + core_stat['sys']
            stat_line.append("{0:.1f}".format(usage))
        output_file.write(','.join(stat_line) + '\n')
    print("Save CPU utilizetion log to {}".format(output_file_name))
    output_file.close()
    return True


def format_cpu_xls_output(cpu_res_arr, output_path, test_name, test_idx, xls_name='', test_report_csv_data=None):
    '''
    @summary: process CPU load results and save it to xls file
    @param cpu_res_arr: dict with parced cpu load test results
    @param output_path: output path to save results
    @param test_name: name of the processed test
    @return True on successfull
    '''
    
    header = ['test','start','end','time','total_time','extra']
    output_file_name = '{}/cpu_usage_report_{}.xlsx'.format(output_path, xls_name)

    if os.path.isfile(output_file_name):
        workbook = openpyxl.load_workbook(filename=output_file_name)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'results'
        sheet.append(header)
        sheet.column_dimensions[get_column_letter(1)].width = 20

    if cpu_res_arr:
        workbook.create_sheet(test_name)

    sheet = workbook.get_sheet_by_name('results')
    test_idx += 1
    if test_report_csv_data:
        for idx in range(len(test_report_csv_data)):
            sheet.cell(row=test_idx, column=idx+2).value = "{0}".format(test_report_csv_data[idx])
    
    if cpu_res_arr:
        sheet.cell(row=test_idx, column=1).value = '=HYPERLINK("#{0}!A1", "{0}")'.format(test_name)
        sheet.cell(row=test_idx, column=1).style = "Hyperlink"
    else:
        sheet.cell(row=test_idx, column=1).value = test_name
        workbook.save(filename=output_file_name)
        workbook.close()
        return True;
    
    sheet = workbook.get_sheet_by_name(test_name)
    print_header = False
    for stat_rec in cpu_res_arr:
        if not print_header:
            core_list = sorted(stat_rec['CPU'].keys())
            header = ['time'] + core_list
            print_header = True
            sheet.append(header)
        stat_line = [stat_rec['time']]
        for core in core_list:
            core_stat = stat_rec['CPU'][core]
            usage = core_stat['usr'] + core_stat['sys']
            stat_line.append(float("{0:.1f}".format(usage)))
        sheet.append(stat_line)

    # creating charts:
    row_count = len(cpu_res_arr)
    column_count = len(core_list)

    chart = LineChart()
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 105
    chart.x_axis.title = "Time sec."
    chart.y_axis.title = "Usage %"
    chart.height = 15
    chart.width = 30

    data = Reference(worksheet=sheet,
                     min_row=1,
                     max_row=row_count + 1,
                     min_col=2,
                     max_col=column_count + 1)

    cats = Reference(worksheet=sheet,
                     min_row=1,
                     max_row=row_count + 1,
                     min_col=1,
                     max_col=1)

    chart.add_data(data, titles_from_data=True)

    series = chart.series[column_count - 1]
    series.graphicalProperties.line.solidFill = "6400A3"
    series.graphicalProperties.line.dashStyle = "sysDash"
    series.graphicalProperties.line.width = 60000  # width in EMUs

    chart.set_categories(cats)
    sheet.add_chart(chart, "l2")

    workbook.save(filename=output_file_name)
    workbook.close()
    return True

def process_test_avg(results_list):
    ''
    result_avg = {'test_duration' : 0, 'overal_duration' : 0, 'extra_info' : 0}
    for result in results_list:
        result_avg['test_duration'] += float(result['test_duration'])
        result_avg['overal_duration'] += float(result['overal_duration'])
        if result['extra_info'] != '':
            result_avg['extra_info'] += float(result['extra_info'])
        else:
            result_avg['extra_info'] = '-'
    result_avg['test_duration'] =  float(result_avg['test_duration'])/len(results_list)
    result_avg['overal_duration'] =  float(result_avg['overal_duration'])/len(results_list)
    if result_avg['extra_info'] != '-':
        result_avg['extra_info']  =  float(result_avg['extra_info'])/len(results_list) 
    return result_avg

def process_results(config):
    '''
    @summary: process all test results passed in config
    @param config: dict with test results
    @return True if no error found
    '''
    res = False
    results = config['test_results']
    out_format = config['format']
    
    test_summary_name = "{}/{}.log".format(config['test_pathname'], BENCHMARK_SUMMARY_FILENAME)
    try:
        test_summary = open(test_summary_name, 'w')
    except IOError as err:
        log.err("I/O error({0}): {1} with log file {2}".format(err.errno,
                                                               err.strerror,
                                                               test_summary_name))
        return False

    test_summary.write("===================================================================================\n")
    test_summary.write("host: '{}' {} benchmark file: '{}'\n". format(config['ip'], config['prefix'], config['bench_plan']))
    test_summary.write("===================================================================================\n")
        
    test_summary_csv_file_name = "{}/{}.csv".format(config['test_pathname'], BENCHMARK_SUMMARY_FILENAME)
    try:
        test_summary_csv = open(test_summary_csv_file_name, 'w')
    except IOError as err:
        log.err("I/O error({0}): {1} with log file {2}".format(err.errno,
                                                               err.strerror,
                                                               test_summary_csv_file_name))
        return False
    header = ['test','start','end','time','total_time','extra']
    test_summary_csv.write(','.join(header) + '\n')
   
    total_test_duration = 0
    total_test_raw_duration = 0  
    test_num = 0
    for test_name, test_res_arr in results.items():
        test_num += 1
        for test_iter in range(len(test_res_arr)):
            test_res = test_res_arr[test_iter]
            test_report_str = TEST_REPORT_FORMAT.format(
                                                test_num=test_num,
                                                test_name=test_name,
                                                test_iter=test_iter,
                                                start=test_res['start_time'].strftime(TEST_REPORT_DATETIME_FORMAT),
                                                end=test_res['stop_time'].strftime(TEST_REPORT_DATETIME_FORMAT),
                                                test_time=test_res['test_duration'],
                                                overal_time=test_res['overal_duration'],
                                                extra_info=test_res['extra_info'])
            test_summary.write(test_report_str + '\n')
            
            test_report_csv_data = ["{}_{}".format(test_name, test_iter),
                                    test_res['start_time'].strftime(TEST_REPORT_DATETIME_FORMAT),
                                    test_res['stop_time'].strftime(TEST_REPORT_DATETIME_FORMAT),
                                    str(test_res['test_duration']),
                                    str(test_res['overal_duration']),
                                    str(test_res['extra_info'])]
            test_summary_csv.write(','.join(test_report_csv_data) + '\n')
            
            total_test_duration += test_res['test_duration']
            total_test_raw_duration += test_res['overal_duration']
            
            cpu_log_arr = None
            cpu_log_filename = ''
            if 'cpu_log' in test_res.keys():
                cpu_log_filename = test_res['cpu_log']
                cpu_log_arr = parse_cpu_log(test_res)
                
            if out_format == 'csv':
                res = format_cpu_csv_output(cpu_log_arr, cpu_log_filename)
            elif out_format == 'xls' and xls_format_ena:
                res = format_cpu_xls_output(cpu_log_arr, 
                                            config['test_pathname'], 
                                            test_name, 
                                            test_num,  
                                            xls_name=test_iter,
                                            test_report_csv_data=test_report_csv_data[1:])
            else:
                print('Unknown output format {}'.format(out_format))
                return False
            
            if res and CPU_USAGE_TEMP_CLEANUP and cpu_log_filename:
                os.remove(cpu_log_filename)

        test_res_avg = process_test_avg(test_res_arr)
        test_report_str = TEST_REPORT_FORMAT.format(
                                                test_num=test_num,
                                                test_name=test_name,
                                                test_iter='avg',
                                                start='-',
                                                end='-',
                                                test_time="%0.2f"%test_res_avg['test_duration'],
                                                overal_time="%0.2f"%test_res_avg['overal_duration'],
                                                extra_info=test_res_avg['extra_info'])
        test_summary.write(test_report_str + '\n')
        
        test_report_csv_data = ["{}_{}".format(test_name, 'avg'),
                                '-',
                                '-',
                                str("%0.2f"%test_res_avg['test_duration']),
                                str("%0.2f"%test_res_avg['overal_duration']),
                                str(test_res_avg['extra_info'])]
        test_summary_csv.write(','.join(test_report_csv_data) + '\n')      
        test_summary.write('\n')
    test_report_summary = TEST_REPORT_SUMARY_FORMAT.format(
        count=config['test_count'],
        passed=config['test_succeed_count'],
        total_time=total_test_duration,
        total_time_raw=total_test_raw_duration
    )
    test_summary.write(test_report_summary + '\n')
    test_summary.close()
    return True


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print "Expected 1 argument\nbench_report.py {cpu_log_file}"
        sys.exit(1)

    cpu_log_file = sys.argv[1]
    output = parse_cpu_log(cpu_log_file)

    print(json.dumps(output, indent=4, sort_keys=True))
    sys.exit(1)
